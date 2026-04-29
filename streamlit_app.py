import os
import streamlit as st
import requests
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime, timedelta
import base64
import re
import json
from pathlib import Path
from typing import Dict, List, Tuple, Optional

# ============================================================================
# CONFIGURATION & SETUP
# ============================================================================

st.set_page_config(
    page_title="Colloscope TSI",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="expanded"
)

CONFIG = {
    "max_groups": 20,
    "max_weeks": 30,
    "max_classes": 2,
    "school_year_start_month": 9,
}

OWNER_CODE = os.getenv("COLLOSCOPE_OWNER_CODE", "debug123")

MATIERES_MAP = {
    'M': "Mathématiques",
    'A': "Anglais",
    'SI': "Sciences de l'Ingénieur",
    'F': "Français",
    'I': "Informatique",
    'P': "Physique",
    'H': "Histoire-Géographie",
    'PHI': "Philosophie",
}

# ============================================================================
# UTILITAIRES
# ============================================================================

def chemin_ressource(chemin_relatif: str) -> str:
    return os.path.join(os.path.abspath("."), chemin_relatif)

def aplatir_liste(liste_imbriquee: List[List[str]]) -> List[str]:
    return [' '.join(sous_liste) for sous_liste in liste_imbriquee]

def extract_date(cell_str: str, year: int) -> Optional[datetime]:
    match = re.search(r'\((\d{2}/\d{2})\)', str(cell_str))
    if match:
        try:
            return datetime.strptime(f"{match.group(1)}/{year}", "%d/%m/%Y")
        except ValueError:
            return None
    return None

def to_naive(dt: datetime) -> datetime:
    if dt.tzinfo is not None:
        return dt.replace(tzinfo=None)
    return dt

def detecter_annee_scolaire_actuelle(date_actuelle: Optional[datetime] = None) -> str:
    if date_actuelle is None:
        date_actuelle = datetime.now()
    annee = date_actuelle.year
    mois = date_actuelle.month
    if mois >= CONFIG["school_year_start_month"]:
        return f"{annee}-{annee + 1}"
    else:
        return f"{annee - 1}-{annee}"

def determiner_matiere(cle: str) -> str:
    for prefix in sorted(MATIERES_MAP.keys(), key=len, reverse=True):
        if cle.startswith(prefix):
            return MATIERES_MAP[prefix]
    return "Non spécifié"

# ============================================================================
# CHARGEMENT DES DONNÉES
# ============================================================================

@st.cache_data(ttl=3600)
def charger_donnees(classe: str, annee_scolaire_str: str) -> Tuple[Dict, Dict, List]:
    annee_numerique = int(annee_scolaire_str.split('-')[0])
    fichier_colloscope = chemin_ressource(f'Colloscope{classe}.xlsx')
    fichier_legende = chemin_ressource(f'Legende{classe}.xlsx')

    try:
        excel_colloscope = load_workbook(fichier_colloscope, data_only=True)
        excel_legende = load_workbook(fichier_legende, data_only=True)

        feuille_colloscope = excel_colloscope.active
        feuille_legende = excel_legende.active

        dictionnaire_donnees: Dict[str, List] = {}
        dates_semaines: List[Optional[datetime]] = []

        # Lecture de la ligne d'en-tête pour les dates (ligne 1)
        premiere_ligne = list(feuille_colloscope.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        for cell_val in premiere_ligne[1:]:  # Ignorer la première colonne (groupes)
            if cell_val is not None:
                date = extract_date(str(cell_val), annee_numerique)
                dates_semaines.append(date)

        # Lecture des données groupes
        for row in feuille_colloscope.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                groupe = str(row[0]).strip()
                donnees_groupe = []
                for cell in row[1:]:
                    if cell is not None:
                        donnees_groupe.append(str(cell).strip())
                    else:
                        donnees_groupe.append("")
                dictionnaire_donnees[groupe] = donnees_groupe

        # Lecture de la légende
        dictionnaire_legende: Dict[str, List] = {}
        for row in feuille_legende.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                cle = str(row[0]).strip()
                valeurs = []
                for cell in row[1:]:
                    if cell is not None:
                        valeurs.append([str(cell).strip()])
                dictionnaire_legende[cle] = valeurs

        return dictionnaire_donnees, dictionnaire_legende, dates_semaines

    except FileNotFoundError as e:
        st.error(f"❌ Fichier manquant : {e}")
        return {}, {}, []
    except Exception as e:
        st.error(f"❌ Erreur lors du chargement : {e}")
        return {}, {}, []

@st.cache_data(ttl=3600)
def obtenir_vacances(zone: str = "C", annee_scolaire_str: Optional[str] = None) -> List[Tuple]:
    if annee_scolaire_str is None:
        annee_scolaire_str = detecter_annee_scolaire_actuelle()

    url = "https://data.education.gouv.fr/api/records/1.0/search/"
    params = {
        "dataset": "fr-en-calendrier-scolaire",
        "rows": 500,
        "refine.zone": f"Zone {zone}",
        "refine.annee_scolaire": annee_scolaire_str,
    }
    vacances = []
    try:
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        for record in data.get("records", []):
            fields = record.get("fields", {})
            start_str = fields.get("start_date")
            end_str = fields.get("end_date")
            if start_str and end_str:
                try:
                    debut = datetime.fromisoformat(start_str)
                    fin = datetime.fromisoformat(end_str)
                    vacances.append((debut, fin))
                except ValueError:
                    pass

        annee_debut_num = int(annee_scolaire_str.split('-')[0])
        annee_fin_num = int(annee_scolaire_str.split('-')[1])
        periode_debut = datetime(annee_debut_num, 9, 1)
        periode_fin = datetime(annee_fin_num, 8, 31)

        vacances = [
            (start, end) for start, end in vacances
            if to_naive(end) >= periode_debut and to_naive(start) <= periode_fin
        ]
    except Exception as e:
        st.warning(f"⚠️ Impossible de récupérer les vacances : {e}")
        vacances = []

    return vacances

# ============================================================================
# PARAMÈTRES
# ============================================================================

CONFIG_FILE = "config.json"

def enregistrer_parametres(groupe: str, semaine: str, classe: str) -> None:
    config = {"groupe": groupe, "semaine": semaine, "classe": classe}
    try:
        with open(CONFIG_FILE, 'w') as f:
            json.dump(config, f)
    except Exception as e:
        st.warning(f"⚠️ Impossible de sauvegarder la config : {e}")

def charger_parametres() -> Tuple[str, str, str]:
    defaults = ("G1", "1", "1")
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f:
                config = json.load(f)
                return (
                    config.get("groupe", defaults[0]),
                    config.get("semaine", defaults[1]),
                    config.get("classe", defaults[2])
                )
        except Exception:
            pass
    return defaults

# ============================================================================
# LOGIQUE MÉTIER
# ============================================================================

def semaine_actuelle(dates_semaines: List[Optional[datetime]], date_actuelle: Optional[datetime] = None) -> int:
    if date_actuelle is None:
        date_actuelle = datetime.now()

    current_weekday = date_actuelle.weekday()
    lundi_actuel = (date_actuelle - timedelta(days=current_weekday)).replace(
        hour=0, minute=0, second=0, microsecond=0
    )

    for i, date_semaine_excel in enumerate(dates_semaines):
        if date_semaine_excel is None:
            continue
        lundi_excel = date_semaine_excel.replace(hour=0, minute=0, second=0, microsecond=0)
        if lundi_excel >= lundi_actuel:
            return i + 1

    return len(dates_semaines) if dates_semaines else 1

def creer_tableau(groupe: str, semaine: str, dictionnaire_donnees: Dict, dictionnaire_legende: Dict) -> List:
    tableau = []
    try:
        semaine_int = int(semaine)
        if groupe not in dictionnaire_donnees:
            st.error(f"❌ Le groupe '{groupe}' n'existe pas dans les données.")
            return tableau

        donnees_groupe = dictionnaire_donnees[groupe]
        if not (1 <= semaine_int <= len(donnees_groupe)):
            st.error(f"❌ La semaine {semaine_int} n'est pas valide pour le groupe '{groupe}'.")
            return tableau

        cle_semaine = donnees_groupe[semaine_int - 1]

        if not cle_semaine:
            return tableau

        # ✅ Supporte plusieurs colles séparées par des espaces (ex: "M10 SI7")
        cles = cle_semaine.split()

        for cle in cles:
            if cle not in dictionnaire_legende:
                st.warning(f"⚠️ Clé '{cle}' introuvable dans la légende.")
                continue

            elements = aplatir_liste(dictionnaire_legende[cle])
            matiere = determiner_matiere(cle)
            elements.append(matiere)
            tableau.append(elements)

    except ValueError:
        st.error("❌ Veuillez entrer une semaine valide (nombre entier).")
    except Exception as e:
        st.error(f"❌ Erreur inattendue lors de la création du tableau : {e}")

    return tableau

def changer_semaine(sens: int) -> None:
    try:
        nouvelle_semaine = int(st.session_state.semaine) + sens
        if 1 <= nouvelle_semaine <= CONFIG["max_weeks"]:
            st.session_state.semaine = str(nouvelle_semaine)
    except (ValueError, TypeError):
        pass

# ============================================================================
# AFFICHAGE
# ============================================================================

def afficher_logo_sidebar() -> None:
    logo_path = "logo_prepa.png"
    if os.path.exists(logo_path):
        try:
            with open(logo_path, "rb") as img_file:
                b64_data = base64.b64encode(img_file.read()).decode()
            html_code = f'''
            <div style="text-align: center; margin-bottom: 30px;">
                <a href="https://sites.google.com/site/cpgetsimarcelsembat/" target="_blank">
                    <img src="data:image/png;base64,{b64_data}" width="120" style="border-radius: 8px;">
                </a>
            </div>
            '''
            st.sidebar.markdown(html_code, unsafe_allow_html=True)
        except Exception as e:
            st.sidebar.warning(f"Logo non disponible : {e}")

def afficher_donnees_colloscope(annee_scolaire_actuelle: str) -> None:
    groupe = st.session_state.groupe
    semaine = st.session_state.semaine
    classe = st.session_state.classe

    enregistrer_parametres(groupe, semaine, classe)

    # Validation groupe
    try:
        numero_groupe = int(groupe[1:])
        if not (1 <= numero_groupe <= CONFIG["max_groups"]):
            st.error(f"❌ Le groupe doit être entre G1 et G{CONFIG['max_groups']}.")
            return
    except (ValueError, IndexError):
        st.error("❌ Format de groupe invalide (ex: G1, G2...).")
        return

    # Validation semaine
    try:
        semaine_int = int(semaine)
        if not (1 <= semaine_int <= CONFIG["max_weeks"]):
            st.error(f"❌ La semaine doit être entre 1 et {CONFIG['max_weeks']}.")
            return
    except ValueError:
        st.error("❌ Semaine invalide.")
        return

    dictionnaire_donnees, dictionnaire_legende, _ = charger_donnees(classe, annee_scolaire_actuelle)

    if not dictionnaire_donnees:
        st.error("❌ Impossible de charger les données du colloscope.")
        return

    donnees = creer_tableau(groupe, semaine, dictionnaire_donnees, dictionnaire_legende)

    if donnees:
        df = pd.DataFrame(donnees, columns=["Professeur", "Jour", "Heure", "Salle", "Matière"])
        df.index = ['' for _ in range(len(df))]
        st.dataframe(df, use_container_width=True, hide_index=True)
    else:
        st.info("ℹ️ Aucune donnée disponible pour cette semaine et ce groupe.")

def afficher_edt_eps() -> None:
    eps_files = ["EPS_page-0001.jpg", "EPS_page-0002.jpg"]
    col1, col2 = st.columns(2)
    for idx, file in enumerate(eps_files):
        col = col1 if idx == 0 else col2
        if os.path.exists(file):
            try:
                col.image(file, caption=f"EDT EPS TSI{idx + 1}", use_container_width=True)
            except Exception as e:
                col.warning(f"Impossible de charger {file}: {e}")
        else:
            col.info(f"📄 {file} non trouvé")

def afficher_dictionnaires_secrets(classe: str, annee_scolaire: str) -> None:
    st.subheader("📚 Contenu des dictionnaires")
    try:
        dictionnaire_donnees, dictionnaire_legende, dates_semaines = charger_donnees(classe, annee_scolaire)

        with st.expander("📊 Dictionnaire de Données"):
            st.json(dictionnaire_donnees)

        with st.expander("📖 Dictionnaire de Légende"):
            st.json(dictionnaire_legende)

        with st.expander("📅 Dates des Semaines"):
            dates_str = [d.strftime("%d/%m/%Y") if d else "None" for d in dates_semaines]
            st.write(dates_str)

    except Exception as e:
        st.error(f"❌ Erreur lors de l'affichage des dictionnaires : {e}")

def afficher_outils_debug(classe: str, annee_scolaire: str) -> None:
    st.subheader("🔧 Outils de Débogage")

    col1, col2, col3 = st.columns(3)

    with col1:
        if st.button("🔄 Recharger données", use_container_width=True, key="btn_reload"):
            st.cache_data.clear()
            st.success("✅ Cache vidé !")
            st.rerun()

    with col2:
        if st.button("🧹 Vider tout le cache", use_container_width=True, key="btn_clear_cache"):
            st.cache_data.clear()
            st.cache_resource.clear()
            st.success("✅ Cache complet vidé !")
            st.rerun()

    with col3:
        if st.button("📋 Afficher état session", use_container_width=True, key="btn_session"):
            st.write("**État du session_state :**")
            st.json(st.session_state.to_dict())

    st.markdown("---")

    col_i1, col_i2, col_i3 = st.columns(3)
    col_i1.metric("Date/Heure", datetime.now().strftime("%d/%m/%Y %H:%M"))
    col_i2.metric("Année Scolaire", annee_scolaire)
    col_i3.metric("Classe Actuelle", classe)

    st.markdown("---")
    st.write("**Test API Vacances**")
    if st.button("📡 Tester récupération vacances", key="btn_vacances"):
        with st.spinner("Récupération en cours..."):
            vacances = obtenir_vacances("C", annee_scolaire)
        if vacances:
            st.success(f"✅ {len(vacances)} période(s) de vacances récupérées.")
            for debut, fin in vacances:
                st.write(f"- {to_naive(debut).strftime('%d/%m/%Y')} → {to_naive(fin).strftime('%d/%m/%Y')}")
        else:
            st.warning("⚠️ Aucune vacance récupérée.")

# ============================================================================
# AUTHENTIFICATION PROPRIÉTAIRE
# ============================================================================

@st.dialog("🔐 Accès Propriétaire")
def dialog_authentification():
    st.write("Entrez le code secret pour accéder aux outils de débogage.")
    code_input = st.text_input(
        "Code secret",
        type="password",
        key="dialog_secret_code",
        placeholder="••••••••"
    )
    col1, col2 = st.columns(2)
    with col1:
        if st.button("✅ Valider", use_container_width=True, key="btn_valider_dialog"):
            if code_input == OWNER_CODE:
                st.session_state["authenticated_owner"] = True
                st.success("✨ Accès accordé !")
                st.rerun()
            else:
                st.error("❌ Code incorrect.")
                st.session_state["authenticated_owner"] = False
    with col2:
        if st.button("❌ Fermer", use_container_width=True, key="btn_fermer_dialog"):
            st.rerun()

# ============================================================================
# APPLICATION PRINCIPALE
# ============================================================================

def main():
    annee_scolaire_actuelle = detecter_annee_scolaire_actuelle()

    groupe_default, semaine_default, classe_default = charger_parametres()

    # Initialisation session_state
    if "groupe" not in st.session_state:
        st.session_state.groupe = groupe_default
    if "semaine" not in st.session_state:
        st.session_state.semaine = semaine_default
    if "classe" not in st.session_state:
        st.session_state.classe = classe_default
    if "authenticated_owner" not in st.session_state:
        st.session_state.authenticated_owner = False

    afficher_logo_sidebar()
    st.sidebar.markdown("---")

    # Onglets dynamiques
    tab_names = ["📅 Colloscope"]
    if st.session_state.get("authenticated_owner", False):
        tab_names.append("🛠️ Outils Propriétaire")

    tabs = st.tabs(tab_names)

    # ===== ONGLET COLLOSCOPE =====
    with tabs[0]:
        st.header("📅 Colloscope TSI")

        with st.expander("🏀 Afficher EDT EPS"):
            afficher_edt_eps()

        st.sidebar.subheader("⚙️ Sélection")

        # Chargement des données pour calculer la semaine actuelle
        try:
            _, _, dates_semaines = charger_donnees(st.session_state.classe, annee_scolaire_actuelle)
            semaines_ecoulees = semaine_actuelle(dates_semaines)
        except Exception:
            semaines_ecoulees = 1

        date_actuelle_str = datetime.now().strftime("%d/%m/%Y")
        st.sidebar.write(f"**Date :** {date_actuelle_str}")
        st.sidebar.write(f"**Semaine actuelle :** {semaines_ecoulees}")
        st.sidebar.write(f"**Année scolaire :** {annee_scolaire_actuelle}")

        # Widgets de sélection
        st.session_state.classe = st.sidebar.selectbox(
            "Classe TSI",
            options=["1", "2"],
            index=["1", "2"].index(st.session_state.classe) if st.session_state.classe in ["1", "2"] else 0,
            key="classe_select"
        )

        st.session_state.groupe = st.sidebar.text_input(
            "Groupe (ex: G1)",
            value=st.session_state.groupe,
            key="groupe_input"
        )

        # Index semaine basé sur session_state (BUG CORRIGÉ)
        semaines_options = [str(i) for i in range(1, CONFIG["max_weeks"] + 1)]
        try:
            semaine_index = semaines_options.index(str(st.session_state.semaine))
        except ValueError:
            semaine_index = semaines_ecoulees - 1

        st.session_state.semaine = st.sidebar.selectbox(
            "Semaine",
            options=semaines_options,
            index=semaine_index,
            key="semaine_select"
        )

        st.sidebar.markdown("---")

        cols = st.sidebar.columns(3)

        if cols[0].button("📋 Afficher", use_container_width=True, key="afficher_btn"):
            st.sidebar.info("⚠️ Vérifiez votre colloscope papier pour éviter les erreurs.", icon="⚠️")
            afficher_donnees_colloscope(annee_scolaire_actuelle)

        if cols[1].button("◀ Préc.", use_container_width=True, key="prev_semaine_btn"):
            changer_semaine(-1)
            st.sidebar.info("⚠️ Vérifiez votre colloscope papier pour éviter les erreurs.", icon="⚠️")
            afficher_donnees_colloscope(annee_scolaire_actuelle)
            st.rerun()

        if cols[2].button("Suiv. ▶", use_container_width=True, key="next_semaine_btn"):
            changer_semaine(1)
            st.sidebar.info("⚠️ Vérifiez votre colloscope papier pour éviter les erreurs.", icon="⚠️")
            afficher_donnees_colloscope(annee_scolaire_actuelle)
            st.rerun()

    # ===== ONGLET OUTILS PROPRIÉTAIRE =====
    if st.session_state.get("authenticated_owner", False) and len(tabs) > 1:
        with tabs[1]:
            st.subheader("🛠️ Outils Propriétaire")

            if st.button("🚪 Déconnexion", key="owner_logout_btn"):
                st.session_state["authenticated_owner"] = False
                st.rerun()

            st.markdown("---")

            debug_tabs = st.tabs(["📚 Dictionnaires", "🔧 Outils de Debug"])

            with debug_tabs[0]:
                if st.button("🔍 Afficher les dictionnaires", key="show_dicts_btn"):
                    afficher_dictionnaires_secrets(st.session_state.classe, annee_scolaire_actuelle)

            with debug_tabs[1]:
                afficher_outils_debug(st.session_state.classe, annee_scolaire_actuelle)

    # Bouton accès propriétaire
    st.sidebar.markdown("<br><br><br>", unsafe_allow_html=True)
    if not st.session_state.get("authenticated_owner", False):
        if st.sidebar.button("🐞", key="owner_access_btn"):
            dialog_authentification()

    # Footer
    st.markdown(
        """
        <div style="margin-top: 40px; padding-top: 20px; border-top: 1px solid #ddd;
                    font-size: 11px; text-align: center; color: #666;">
            Fait par <strong>BERRY Mael</strong>, avec l'aide de SOUVELAIN Gauthier,
            ChatGPT, Gemini et DAMBRY Paul<br/>
            <em>Version améliorée 2024</em>
        </div>
        """,
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
